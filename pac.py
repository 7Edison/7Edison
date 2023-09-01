import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
import sqlite3


# #连接数据库
conn = sqlite3.connect('my_database.db')
cur = conn.cursor()
# #创建数据表
cur.execute('''
    CREATE TABLE IF NOT EXISTS infos (
        id INTEGER PRIMARY KEY,
        lab TEXT,
        address TEXT,
        room TEXT,
        area TEXT,
        number TEXT,
        place TEXT,
        result TEXT,
        fenqu TEXT,
        url TEXT)
''')

#设置请求头
headers = {
"Accept":"text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
"cookie":"SECKEY_ABVK=2DaNir5uzWWbUcDjQU/9H+xSHEBl+/flIWQd/KZRzv8%3D; BMAP_SECKEY=SYb8czaOW960nA2EmyMlvu0m9CmKq5FrEIFcRbUl0cL31TRG3R0RQIzwpFsXOTK_gMjfTbTIRvi22zb3kfjbPt1tCSCS3MyuFQHW4wqLbPwC0oJG0_7l9ukAtkyRPCMtfdk43GDIcAK1254PT3BioGNRUjh5VThplSA1a9XUbr0OTP-aIhZ8U3YPex9iU_RW; lps=https%3A%2F%2Fsh.zu.anjuke.com%2Ffangyuan%2F%7C; cmctid=2; aQQ_ajkguid=A834FA75-6FE4-A307-79CB-3761B084AAF3; id58=CrIekmTxXVGz+3EZHpFBAg==; wmda_visited_projects=%3B6289197098934; id58=CrIekmTxXVK9C3EfHpf+Ag==; wmda_uuid=c736f2c5d2619b25f9876ef288a7ac89; wmda_new_uuid=1; sessid=073369FB-F36E-69F8-49E4-F03690ABD3B4; obtain_by=2; twe=2; 58tj_uuid=d3b979fa-dc18-44a8-8cb7-3123e8e74083; new_session=1; init_refer=https%253A%252F%252Fcn.bing.com%252F; new_uv=1; als=0; ajk-appVersion=; fzq_h=e32b320d813bae0a7695094e2bf1a119_1693546916651_bb206a7262bb4ce290a4ccbdc6bced86_3060683597; ctid=11; wmda_session_id_6289197098934=1693546936332-9e2037b2-b931-3e49; xxzl_cid=cb3ab709de684ccf95b185262008735b; xxzl_deviceid=u6Zhg+rvSr1Wqu90FUEkDUd80fzl9k3Jw7XogHK09i/sB4xxvEO7OmAHwZZPhHkH",
"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36 Edg/116.0.1938.62"
}
#暂时存储数据
all_data=[]
#定义爬取信息的函数
def get_data(url,qu):
    # 向url发起请求
    response = requests.get(url, headers=headers)
    # 获得响应内容
    # print(response.text)
    html_content = response.content
    # 用BeautifulSoup对html进行解析
    soup = BeautifulSoup(html_content, "html.parser")
    #获取标签为div，class为zu-itemmod里面的内容
    html = soup.find_all('div', class_='zu-itemmod')
    for element in html:
        data={}
        lab = element.img['alt']  #标签
        # print(f"lable:{lab}")
        data["lab"]=lab
        address = element.address.get_text(strip=True) #地址
        data["address"] = address
        # print(f"address:{address }")
        place=element.find("div",class_="zu-side").get_text(strip=True)  #价格
        data["place"] = place.replace("元/月","")
        # print(f"place:{place}")
        keyword=element.find("p",class_='details-item bot-tag').get_text(strip=False)  #关键词
        key = keyword.strip().splitlines()
        result = ','.join(key)
        data["result"] = result
        msg1=element.find("p",class_='details-item tag')
        msg=msg1.get_text()
        data_list = msg.split()
        msgs=data_list[0].split("|")
        print(f"number：{data_list[0]}")
        data["room"] = msgs[0]    #房间
        data["area"] = msgs[1].replace("平米","")      #面积
        data["number"] = msgs[2]      #楼层
        url=element.a["href"]   #网址
        data["url"] = url
        data["fenqu"]=qu
        print(f"data:{data}")
             #所属分区
        all_data.append(data)

    return all_data

#需要爬取的区
dirs={"pudong":"浦东","minhang":"闵行","songjiang":"松江","jiading":"嘉定","xuhui":"徐汇","qingpu":"青浦","jingan":"静安","huangpu":"黄埔","hongkou":"虹口"}
# dirs={"pudong":"浦东","huangpu":"黄埔","hongkou":"虹口"}
dir_url=list(dirs.keys())
#进行爬取
for i in dir_url:
    for j in range(1,20):
        url=f"https://sh.zu.anjuke.com/fangyuan/{i}/p{j}/?pi=360-cpchz-sh-hexin-shzf1&utm_source=360-qg&kwid=39710887124&utm_term=%E4%B8%8A%E6%B5%B7+%E7%A7%9F%E6%88%BF"
        get_data(url,dirs[i])

# all_data是一个列表，里面每个数据是字典，如[{'lab': 'a', 'address': 'b'},{'lab': 'c', 'address': 'd'}.....]
# 可以遍历列表，然后从字典依次取值存入数据库

# 将数据插入数据库
for item in all_data:
    cur.execute('INSERT INTO infos (lab, address, room, area, number, place, result, fenqu, url) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)',
                   (item['lab'], item['address'], item['room'], item['area'], item['number'], item['place'], item['result'], item['fenqu'], item['url']))

conn.commit()     #提交
conn.close()      #关闭



#将爬取结果写入excle表格
workbook = Workbook()
# 获取默认的工作表
worksheet = workbook.active
# 列标题
headers = ["序号",'简介', '地址',"房间","面积(平米)","层数", '价格(元/每月)','标签',"区","网址"]
for col_num, header in enumerate(headers, 1):
    worksheet.cell(row=1, column=col_num, value=header)
for row_num, data in enumerate(all_data, 2):  # 从第二行开始写入数据
    worksheet.cell(row=row_num, column=1, value=row_num - 1)
    worksheet.cell(row=row_num, column=2, value=data.get('lab'))
    worksheet.cell(row=row_num, column=3, value=data.get('address'))
    worksheet.cell(row=row_num, column=4, value=data.get('room'))
    worksheet.cell(row=row_num, column=5, value=data.get('area'))
    worksheet.cell(row=row_num, column=6, value=data.get('number'))
    worksheet.cell(row=row_num, column=7, value=data.get('place'))
    worksheet.cell(row=row_num, column=8, value=data.get('result'))
    worksheet.cell(row=row_num, column=9, value=data.get('fenqu'))
    worksheet.cell(row=row_num, column=10, value=data.get('url'))

# 保存工作簿
workbook.save('data3.xlsx')



